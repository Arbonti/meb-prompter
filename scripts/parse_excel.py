import openpyxl
import xlrd
import json
import os

data_dir = r"C:\Users\alpte_gwx6ike\OneDrive\Masaüstü\MEB PROMPTER\data"
out_dir = r"C:\Users\alpte_gwx6ike\OneDrive\Masaüstü\MEB PROMPTER"

def parse_tablo_xlsx(path):
    """Parse genel tablo xlsx (özet istatistikler)"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(c).strip() if c is not None else "" for c in row]
                if any(r for r in row_data):
                    rows.append(row_data)
            result[sheet_name] = rows
        return result
    except Exception as e:
        print(f"Error parsing {path}: {e}")
        return {}

def parse_detail_xls(path):
    """Parse detail xls dosyaları (okul okul liste)"""
    try:
        wb = xlrd.open_workbook(path)
        result = {}
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            for i in range(ws.nrows):
                row = [str(ws.row_values(i)[j]).strip() if ws.row_values(i)[j] != '' else '' for j in range(ws.ncols)]
                if any(r for r in row):
                    rows.append(row)
            result[sheet_name] = rows
        return result
    except Exception as e:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            result = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_data = [str(c).strip() if c is not None else "" for c in row]
                    if any(r for r in row_data):
                        rows.append(row_data)
                result[sheet_name] = rows
            return result
        except Exception as e2:
            print(f"Error: {e2}")
            return {}

def to_num(v):
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except:
        return v


data = {
    "2021-2022": {},
    "2020-2021": {}
}


print("Parsing 2021-2022 TABLO...")
tablo_2122 = parse_tablo_xlsx(os.path.join(data_dir, "2021-2022_TABLO.xlsx"))
data["2021-2022"]["genel"] = tablo_2122


for key, fname in [("okuloncesi","okuloncesi_2122.xlsx"),("ilkokul","ilkokul_2122.xls"),("ortaokul","ortaokul_2122.xls"),("lise","lise_2122.xls")]:
    fpath = os.path.join(data_dir, fname)
    print(f"Parsing 2021-2022 {key}...")
    detail = parse_detail_xls(fpath)
    data["2021-2022"][key] = detail


print("Parsing 2020-2021 TABLO...")
tablo_2021 = parse_tablo_xlsx(os.path.join(data_dir, "2020-2021_TABLO.xlsx"))
data["2020-2021"]["genel"] = tablo_2021


for key, fname in [("okuloncesi","okuloncesi_2021.xlsx"),("ilkokul","ilkokul_2021.xls"),("ortaokul","ortaokul_2021.xls"),("lise","lise_2021.xls")]:
    fpath = os.path.join(data_dir, fname)
    print(f"Parsing 2020-2021 {key}...")
    detail = parse_detail_xls(fpath)
    data["2020-2021"][key] = detail


out_path = os.path.join(out_dir, "data_raw.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\nSaved to {out_path}")
print("Keys per year:")
for yr in data:
    print(f"  {yr}: {list(data[yr].keys())}")
    for k in data[yr]:
        if isinstance(data[yr][k], dict):
            for sheet in data[yr][k]:
                rows = data[yr][k][sheet]
                print(f"    {k}/{sheet}: {len(rows)} rows")
                if rows:
                    print(f"      First row: {rows[0][:8]}")
                    if len(rows) > 1:
                        print(f"      2nd row:   {rows[1][:8]}")
                    if len(rows) > 2:
                        print(f"      3rd row:   {rows[2][:8]}")
